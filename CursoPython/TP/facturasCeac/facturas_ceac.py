#Se importan los modulos necesarios
import os
from datetime import datetime
import locale
import re
import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side

#Se establece el directorio actual
directorio_actual = os.path.dirname(os.path.abspath(__file__))
os.chdir(directorio_actual)
print(f"El directorio actual es {os.getcwd()}")
#Se crea carpeta con el mes actual

locale.setlocale(locale.LC_ALL, ("es_ES", "UTF-8"))
mes = str(datetime.now().strftime("%B %Y")).upper()
nombre_carpeta = f"Comprobantes CEAC {mes}"

#Se crea carpeta nueva
directorio_carpeta = os.path.join(directorio_actual, nombre_carpeta)
carpeta_nueva_pdfs = os.path.join(nombre_carpeta, " PDFs")
os.makedirs(carpeta_nueva_pdfs, exist_ok=True)

#leer mensaje
print(f"os.getcwd() {os.getcwd()}")
directorio_mensaje = os.path.join(os.getcwd(), "mensaje.txt")
with open(directorio_mensaje, "r") as archivo:
    mensaje = archivo.read()
#trabajando con el mensaje
patrones_mensajes = re.compile(
    r'(?P<tipo>Agua|Cable|Energia|Internet)\s+Venc\.: (?P<vencimiento>\d{2}/\d{2}/\d{4}) \$'
    r'(?P<saldo>[\d,.]+)\s+(?P<pdf_url>https://levalle\.micoop\.com\.ar/impresion\?\w+)\s+Pagar\s+'
    r'(?P<pay_url>https://levalle\.micoop\.com\.ar/comprobantes\?[\w=&\[\]]+)',
    re.MULTILINE
)

#guardamos en listas los datos que nos interesan

pdfs = []
for match in patrones_mensajes.finditer(mensaje):
    pdfs.append(match.groupdict())

links_pdf = [c["pdf_url"] for c in pdfs]
#print(links_pdf)
links_pago = [c["pay_url"] for c in pdfs]
#print(links_pago)
saldos = [c["saldo"] for c in pdfs]
#print(saldos)
vencimiento = [c["vencimiento"] for c in pdfs]
#print(vencimiento)
tipo = [c["tipo"] for c in pdfs]
#print(tipo)

#Saldo total
saldo_total = sum([float(s.replace(".","").replace(",", ".")) for s in saldos])
saldo_total = round(saldo_total, 2)
print(f"Saldo total: {saldo_total}")

#Descargar PDF
def descargar_pdf(url, tipo, vencimiento, carpeta_destino,i):
    nombre_archivo = f"Comprobante N° {i+1} Mes {mes} {tipo} - Vencimiento {vencimiento}.pdf"
    ruta_archivo = os.path.join(carpeta_destino, nombre_archivo.replace("/", "-"))
    with open(ruta_archivo, 'wb') as archivo:
        response = requests.get(url)
        archivo.write(response.content)
    print(f"Archivo PDF descargado: {nombre_archivo}")

# Descargar cada PDF
for i, pdf_url in enumerate(links_pdf):
    tipo_archivo = tipo[i]
    fecha_vencimiento = vencimiento[i]
    descargar_pdf.contador = 1  
    descargar_pdf(pdf_url, tipo_archivo, fecha_vencimiento, carpeta_nueva_pdfs, i)

# Crear un archivo Excel con los links
data = {
    'Factura': [f'factura_{i}.pdf' for i in range(len(pdfs))],
    'Vencimiento': [c["vencimiento"] for c in pdfs],
    'Saldo a Pagar': [float(c['saldo'].replace('.', '').replace(',', '.')) for c in pdfs],
    'Link de Pago': [c['pay_url'] for c in pdfs]
}
df = pd.DataFrame(data)
df.loc['Total'] = df.sum(numeric_only=True)
df.at['Total', 'Factura'] = 'Total'
df.at['Total', 'Vencimiento'] = ''
df.at['Total', 'Link de Pago'] = 'https://levalle.micoop.com.ar/comprobantes?user=3462'

nombre_excel = os.path.join(f"Resumen Comprobantes CEAC {mes}.xlsx")
try:
    # Escribir el DataFrame en un archivo Excel
    df.to_excel(nombre_excel, index=False)
    print(f"Archivo Excel guardado exitosamente en {nombre_excel}")

    # Leer el archivo Excel recién creado
    wb = load_workbook(nombre_excel)
    ws = wb.active

    # Ajustar el ancho de las columnas
    ws.column_dimensions['A'].width = 15  # Ancho de la columna Factura
    ws.column_dimensions['B'].width = 20  # Ancho de la columna Vencimiento
    ws.column_dimensions['C'].width = 30  # Ancho de la columna Saldo a Pagar
    ws.column_dimensions['D'].width = 25  # Ancho de la columna Link de Pago

    # Convertir la columna D en hipervínculos
    for row_idx in range(2, len(df) + 2):
        cell = ws[f'D{row_idx}']
        cell.value = f'=HYPERLINK("{cell.value}", "{cell.value}")'
    # Modificar el formato de la columna C
    for row in ws.iter_rows(min_row=2, min_col=3, max_row=ws.max_row, max_col=3):
        for cell in row:
            cell.font = Font(size=20, bold=True)
    # Subrayar toda la última fila agregada
    last_row = ws.max_row
    for cell in ws[f'A{last_row}':f'D{last_row}']:
        for c in cell:
            c.border = Border(bottom=Side(border_style='thin'))
    # Guardar los cambios
    wb.save(nombre_excel)

    print(f"Ancho de las columnas ajustado y archivo Excel guardado exitosamente en {nombre_excel}")

except Exception as e:
    print(f"Error al guardar el archivo Excel: {e}")